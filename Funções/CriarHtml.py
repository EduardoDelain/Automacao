import openpyxl as xl
from template import PegarHtml
from GerarLog.Log import log
from Funções.EnviarEmail import envia_email

def criar_html(linha_pedidos_por_cliente: dict, aba: "planilha") -> None:
    """
    Recebe as linhas com os pedidos de cada cliente e cria o html necessario para o relatorio
    :param linha_pedidos_por_cliente: Dicionario que contem as linha do excel com pedido de cada cliente
    :param aba: worksheet do excel
    :return: None
    """
    for key in linha_pedidos_por_cliente:
        log.mensagem(f'Criando o html para o cliente {key}')
        valor_total = 0
        numero_vendas = 0
        html = PegarHtml.header
        html += f"""<h1>{key}</h1>"""
        html += PegarHtml.html_inicio
        log.mensagem(f'Calculando os dados a serem exibidos no email')
        for value in linha_pedidos_por_cliente[key]:
            status = aba.cell(row=value, column=5).value
            valor_desconto = aba.cell(row=value, column=9).value
            valor_pago = aba.cell(row=value, column=10).value
            forma_pagamento = aba.cell(row=value, column=11).value
            if status == 'Aprovado':
                numero_vendas += 1
                valor_total += valor_pago
        html += f"""
            <div>
                <span>Faturamento</span>
                <h1>R$ {round(valor_total, 2)}</h1>
            </div>
            <div>
                <span>Numero de vendas</span>
                <h1>{numero_vendas}</h1>
            </div>
        """

        envia_email(html)
    log.mensagem('-----------------EXECUÇÃO FINALIZADA--------------------')