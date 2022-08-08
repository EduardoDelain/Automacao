import openpyxl as xl
from GerarLog.Log import log
from typing import List, Tuple, Dict
import smtplib
import email.message
import glob

def selecionar_tabela():
    """
    Seleciona a tabela e worksheet a ser usado na aplicação
    :return: worksheet a ser usado na aplicação
    """
    log.mensagem('---------------------COMEÇANDO A EXECUÇÃO--------------------------------')
    log.mensagem('Lendo a planilha.......')
    try:
        caminho = input('Qual o caminho da planilha que deseja ler?')
        nome_aba = input('Qual o nome da aba da planilha que deseja ler?')
        planilha = xl.load_workbook(filename=caminho, data_only=True)
        aba = planilha[nome_aba]
    except:
        log.mensagem('O caminho ou a aba que foi colocava estava incorreta')
        log.mensagem('-----------------EXECUÇÃO FINALIZADA--------------------')
    log.mensagem('Tabela selecionada')
    return aba

def selecionar_todos_clientes(aba: "planilha") -> dict:
    """
    Função que cria o dicionario "Clientes" e armazena nele o ID de cada cliente e a linha de cada compra desse cliente
    :param aba: worksheet do excel
    :return: Dicionario cliente com as Informações
    """
    linha_pedidos_por_cliente = {}
    for linha in range(2, aba.max_row):
        nome = aba.cell(row=linha, column=1).value
        aprovacao = aba.cell(row=linha, column=5).value
        if nome not in linha_pedidos_por_cliente:
            linha_pedidos_por_cliente[nome] = []
            linha_pedidos_por_cliente[nome].append(linha)
        else:
            linha_pedidos_por_cliente[nome].append(linha)
    log.mensagem('Clientes selecionados')
    return linha_pedidos_por_cliente

